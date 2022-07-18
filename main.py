import datetime
import http.client as httplib
import json
import os
import re
import time
import urllib.parse
from http.cookies import SimpleCookie
from pathlib import Path

import openpyxl
import requests
from PIL import Image, UnidentifiedImageError
from bs4 import BeautifulSoup
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console
from urllib3.exceptions import ReadTimeoutError


def have_internet() -> bool:
    """
    Checks if internet connection is ON

    :return: True if it can read headers of google DNS server (8.8.8.8)
    """
    conn = httplib.HTTPSConnection("8.8.8.8", timeout=5)
    try:
        conn.request("HEAD", "/")
        return True
    except Exception:
        return False
    finally:
        conn.close()


def get_excel_sheet(filename: str) -> Worksheet:
    """
    Get active worksheet

    :param filename: path to the file
    :return: worksheet
    """
    wb = openpyxl.load_workbook(filename, read_only=True)
    return wb.active


def strfdelta(tdelta, fmt):
    """
    Format timedelta

    :param tdelta: timedelta
    :param fmt: format i.e. {hours}:{minutes}:{seconds}
    """
    d = {"days": tdelta.days}
    d["hours"], rem = divmod(tdelta.seconds, 3600)
    d["minutes"], d["seconds"] = divmod(rem, 60)
    return fmt.format(**d)


def print_status(start_time: datetime, current_position: int, total_positions: int, passed_positions: int = 0) -> str:
    """
    Count current position and calculate time till all files will be saved, average time per file

    :param passed_positions: how many files was already processed before (exclude them from time per file calc)
    :param start_time: datetime when processing started
    :param current_position: current row in excel sheet
    :param total_positions: total rows in Excel sheet
    """
    # remove how many files that already been downloaded
    current_position += 1

    # calculate time difference
    time_diff = datetime.datetime.now() - start_time
    speed = time_diff.seconds / (current_position - passed_positions)
    time_left = (total_positions - current_position) * speed
    formatted_time_left = strfdelta(datetime.timedelta(seconds=time_left), "{hours}:{minutes}:{seconds}")

    return f"{current_position} out of {total_positions}." \
           f" estimated time left {formatted_time_left}. Time per single file in avg {round(speed, 2)} sec"


def save_image(url_hires: str, url_lowres: str, file_name: str, file_format: str = 'WEBP') -> bool:
    """
    Trying to save image from given URLs

    :param url_hires: link to HIGH resolution image
    :param url_lowres: link to LOW resolution image
    :param file_name: name file for saving
    :param file_format: i.e.: PNG, JPG, WEBP...
    :raises UnidentifiedImageError: if PIL cannot read image
    :return: True if saved successfully, else - False
    """
    if url_hires.startswith('//'):
        url_hires = 'https:' + url_hires

    if url_lowres.startswith('//'):
        url_lowres = 'https:' + url_lowres

    try:
        if url_lowres == url_hires:
            img = Image.open(requests.get(url_lowres, stream=True, timeout=15).raw)
            img.save(f'img/lo-res/{file_name}.{file_format}', format=file_format)
            img.save(f'img/hi-res/{file_name}.{file_format}', format=file_format)
        else:
            img = Image.open(requests.get(url_lowres, stream=True, timeout=15).raw)
            img.save(f'img/lo-res/{file_name}.{file_format}', format=file_format)
            img = Image.open(requests.get(url_hires, stream=True, timeout=15).raw)
            img.save(f'img/hi-res/{file_name}.{file_format}', format=file_format)
    except UnidentifiedImageError:
        return False
    except Exception:
        return False
    return True


def create_img_folders():
    """
    Creates folders for images to save if they not exist
    """
    path = 'img/hi-res'
    os.makedirs(path, exist_ok=True)
    path = 'img/lo-res'
    os.makedirs(path, exist_ok=True)


def get_cookies() -> dict:
    """
    Get cookies from env or user input if env not found
    """
    raw_cookie = os.getenv('COOKIES')
    if raw_cookie is None:
        raw_cookie = input('Enter cookies: ')
    cookie = SimpleCookie()
    cookie.load(raw_cookie)  # read cookie from env
    return {k: v.value for k, v in cookie.items()}


def main():
    header = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
                            "(KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"}

    # create folders
    create_img_folders()

    # console
    console = Console()

    # get filename/file path
    file = input('Enter path to excel file: ')
    while not Path.is_file(Path(file)):
        # console.log('[bold red]ERROR: File not found, try again...')
        file = input('File doesnt exist, please enter correct path to excel file: ')


    image_format = 'webp'
    image_format = input('What format you want to save images? (webp, jpg, png, gif...): ')
    while '.' + image_format.lower() not in Image.registered_extensions().keys():
        image_format = input('Unsupported format, please enter another one (webp, jpg, png, gif...): ')

    # delay to not get banned by yandex (less than 2 sec not recommended)
    request_delay = None
    while not request_delay:
        try:
            request_delay = int(input('Choose delay between requests to yandex (recommend 2 sec to avoid ban): '))
        except ValueError:
            pass

    # column in Excel file to read names of songs from
    column_index = 0

    # read cookies
    cookies = get_cookies()

    # open Excel file
    console.log('[bold green]Reading excel file', highlight=False)
    sheet = get_excel_sheet(file)

    # takes current time on start to measure average time to complete operations
    start_time = datetime.datetime.now()

    # counter for files to exclude from calculation time spent for each file to download
    amount_existed_files = 0
    with console.status('[bold blue]Downloading...') as status:
        for index, row in enumerate(sheet.iter_rows()):
            status.update('[bold blue]' + print_status(start_time, index, sheet.max_row, amount_existed_files))

            cell = row[column_index].value

            # check if file already exist to not re download it
            if os.path.exists(f'img/hi-res/{cell}.{image_format}'):
                amount_existed_files += 1
                continue

            # search text for yandex
            search_query = re.sub(r"[\(\[].*?[\)\]]", "", cell).strip() + ' обложка'

            # check if we have connection, pause if no connection
            no_conn_counter = 0
            while not have_internet():
                no_conn_counter += 1
                console.log(f'\r[bold red]INTERNET CONNECTION OFF:[/bold red] for {no_conn_counter * 10} sec',
                            highlight=False,
                            end='')
                time.sleep(10)

            # main. trying to get response from yandex
            try:
                request = requests.get(
                    f'https://yandex.ru/images/search?iorient=square&text='
                    f'{urllib.parse.quote(search_query)}',
                    headers=header, cookies=cookies, timeout=10)
            except ReadTimeoutError:
                console.log(f'[bold red]ERROR:[/bold red] URL TIMEOUT for {cell}', highlight=False, end='')
                continue
            except Exception as e:
                console.log(f'[bold red]ERROR:[/bold red] {e}', highlight=False, end='')
                continue
            finally:
                # delay before next request
                time.sleep(request_delay)

            if not request:
                console.log(f'[bold red]ERROR:[/bold red] request is NONE for {cell}', highlight=False, end='')
                continue

            # parsing content
            soup = BeautifulSoup(request.content, "html.parser")

            item = soup.find(class_='serp-item_type_search')
            try:
                image_json = json.loads(item['data-bem'])
            except Exception:
                console.log(f'[bold red]ERROR:[/bold red] cant resolve data-bem for {cell}', highlight=False, end='')
                continue

            hi_res_url = image_json['serp-item']['preview'][0]['url']

            item = soup.find(class_='serp-item__link')
            low_res_url = item.find('img')['src']

            if low_res_url == '':
                console.log(f'[bold red]ERROR:[/bold red] URL was empty for {cell}', highlight=False, end='')
                continue

            if save_image(hi_res_url, low_res_url, row[column_index].value, image_format):
                console.log(f'[bold green]SUCCESS:[/bold green] {cell}', highlight=False, end='')
            else:
                console.log(f'[bold red]FAILED:[/bold red] cannot save file for {cell}', highlight=False, end='')


if __name__ == '__main__':
    main()
